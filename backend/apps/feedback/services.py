from django.db import transaction
from django.db.models import Avg
from rest_framework.exceptions import ValidationError, NotFound, PermissionDenied

from apps.audit.models import AuditLog
from apps.reviewer_workflow.models import ReviewerTask
from apps.review_cycles.models import ReviewCycle, CycleParticipant, TemplateQuestion
from .models import FeedbackResponse, FeedbackAnswer, AggregatedResult


# ─── Submit Feedback ──────────────────────────────────────────────────────────

def submit_feedback(task_id, user, answers):
    try:
        task = ReviewerTask.objects.select_related(
            'cycle__template', 'reviewee'
        ).get(id=task_id)
    except ReviewerTask.DoesNotExist:
        raise NotFound('Task not found')

    if task.reviewer != user:
        raise PermissionDenied('Access denied')

    if task.status not in ['CREATED', 'PENDING', 'IN_PROGRESS']:
        raise ValidationError('Task is already submitted or locked')

    if task.cycle.state != 'ACTIVE':
        raise ValidationError('Cycle is not in active state')

    if not answers:
        raise ValidationError('answers array is required')

    # Load all questions for this template (used for required + scale validation)
    all_questions = {
        str(q.id): q for q in TemplateQuestion.objects.filter(
            section__template=task.cycle.template
        )
    }

    # Validate required questions are answered
    required_q_ids = {qid for qid, q in all_questions.items() if q.is_required}
    answered_q_ids = {str(a['question_id']) for a in answers}
    missing = required_q_ids - answered_q_ids
    if missing:
        raise ValidationError(f'Missing required answers for {len(missing)} question(s)')

    # Validate answers are for questions that belong to this template
    unknown = answered_q_ids - set(all_questions.keys())
    if unknown:
        raise ValidationError(f'Answer contains unknown question IDs')

    # Validate rating values are within each question's defined scale
    scale_errors = []
    for a in answers:
        qid = str(a['question_id'])
        q = all_questions.get(qid)
        if q and q.type == 'RATING' and a.get('rating_value') is not None:
            rv = float(a['rating_value'])
            if q.rating_scale_min is not None and rv < q.rating_scale_min:
                scale_errors.append(
                    f'Rating {rv} is below minimum {q.rating_scale_min} for question "{q.question_text[:50]}"'
                )
            if q.rating_scale_max is not None and rv > q.rating_scale_max:
                scale_errors.append(
                    f'Rating {rv} exceeds maximum {q.rating_scale_max} for question "{q.question_text[:50]}"'
                )
    if scale_errors:
        raise ValidationError(scale_errors)

    with transaction.atomic():
        # Remove any existing response (re-submission guard)
        FeedbackResponse.objects.filter(task=task).delete()

        response = FeedbackResponse.objects.create(task=task, submitted_by=user)

        FeedbackAnswer.objects.bulk_create([
            FeedbackAnswer(
                response=response,
                question_id=a['question_id'],
                rating_value=a.get('rating_value'),
                text_value=a.get('text_value') or '',
            )
            for a in answers
        ])

        task.status        = 'SUBMITTED'
        task.draft_answers = None
        task.save(update_fields=['status', 'draft_answers', 'updated_at'])

    AuditLog.log(actor=user, action='SUBMIT_FEEDBACK',
                 entity_type='reviewer_task', entity_id=task_id,
                 new_value={
                     'reviewer': user.get_full_name(),
                     'reviewee': task.reviewee.get_full_name(),
                     'task_type': task.reviewer_type,
                     'cycle': task.cycle.name,
                 })

    return {'response_id': str(response.id), 'task_id': str(task_id)}


# ─── Aggregation Pipeline ─────────────────────────────────────────────────────

def aggregate_cycle(cycle):
    """
    Calculate overall / self / manager / peer scores for every participant.
    Idempotent — safe to run multiple times (uses update_or_create).
    Uses a single aggregation query per reviewee instead of N+1 per reviewer_type.
    """
    from django.db.models import Avg, Q

    participants = CycleParticipant.objects.filter(cycle=cycle).select_related('user')

    for participant in participants:
        reviewee = participant.user

        # One query: get all averages for this reviewee broken down by reviewer_type
        scores = (
            FeedbackAnswer.objects
            .filter(
                response__task__cycle=cycle,
                response__task__reviewee=reviewee,
                response__task__status='SUBMITTED',
                rating_value__isnull=False,
            )
            .aggregate(
                overall=Avg('rating_value'),
                self_score=Avg('rating_value', filter=Q(response__task__reviewer_type='SELF')),
                manager_score=Avg('rating_value', filter=Q(response__task__reviewer_type='MANAGER')),
                peer_score=Avg('rating_value', filter=Q(response__task__reviewer_type='PEER')),
            )
        )

        AggregatedResult.objects.update_or_create(
            cycle=cycle,
            reviewee=reviewee,
            defaults={
                'overall_score': scores['overall'],
                'self_score':    scores['self_score'],
                'manager_score': scores['manager_score'],
                'peer_score':    scores['peer_score'],
            }
        )


# ─── Report Helpers ───────────────────────────────────────────────────────────

def _get_feedback_sections(cycle, reviewee, viewer_role, viewer):
    """
    Builds the detailed feedback sections for a report.
    Applies anonymity rules — hides reviewer identity for ANONYMOUS tasks.
    """
    tasks = ReviewerTask.objects.filter(
        cycle=cycle, reviewee=reviewee, status='SUBMITTED'
    ).select_related('reviewer')

    sections = []
    for task in tasks:
        response = FeedbackResponse.objects.filter(task=task).first()
        if not response:
            continue

        answers = FeedbackAnswer.objects.filter(response=response).select_related('question')

        # Determine identity visibility based on anonymity mode:
        # TRANSPARENT  → everyone sees identity
        # SEMI_ANONYMOUS → HR_ADMIN/MANAGER/SUPER_ADMIN see identity; employee does not
        # ANONYMOUS    → only SUPER_ADMIN sees identity (preserves anonymity for all others)
        show_identity = (
            task.anonymity_mode == 'TRANSPARENT'
            or viewer_role == 'SUPER_ADMIN'
            or (task.anonymity_mode == 'SEMI_ANONYMOUS' and viewer_role in ['HR_ADMIN', 'MANAGER'])
            or task.reviewer == viewer
        )

        sections.append({
            'reviewer_type': task.reviewer_type,
            'anonymity_mode': task.anonymity_mode,
            'hidden': not show_identity,
            'identity': {
                'id':         str(task.reviewer.id),
                'first_name': task.reviewer.first_name,
                'last_name':  task.reviewer.last_name,
            } if show_identity else None,
            'answers': [
                {
                    'question_id':   str(a.question_id),
                    'question_text': a.question.question_text,
                    'question_type': a.question.type,
                    'rating_value':  float(a.rating_value) if a.rating_value is not None else None,
                    'text_value':    a.text_value,
                }
                for a in answers
            ]
        })

    return sections


# ─── My Report (Employee) ─────────────────────────────────────────────────────

def get_my_report(cycle_id, user):
    try:
        cycle = ReviewCycle.objects.get(id=cycle_id)
    except ReviewCycle.DoesNotExist:
        raise NotFound('Cycle not found')

    if cycle.state not in ['RESULTS_RELEASED', 'ARCHIVED']:
        raise PermissionDenied('Results are not yet released')

    if not CycleParticipant.objects.filter(cycle=cycle, user=user).exists():
        raise PermissionDenied('You are not a participant in this cycle')

    try:
        aggregated = AggregatedResult.objects.get(cycle=cycle, reviewee=user)
    except AggregatedResult.DoesNotExist:
        aggregated = None

    sections = _get_feedback_sections(cycle, user, user.role, user)

    return {
        'cycle_id':      str(cycle_id),
        'cycle_name':    cycle.name,
        'reviewee':      {'id': str(user.id), 'name': user.get_full_name()},
        'overall_score': float(aggregated.overall_score) if aggregated and aggregated.overall_score is not None else None,
        'self_score':    float(aggregated.self_score)    if aggregated and aggregated.self_score    is not None else None,
        'manager_score': float(aggregated.manager_score) if aggregated and aggregated.manager_score is not None else None,
        'peer_score':    float(aggregated.peer_score)    if aggregated and aggregated.peer_score    is not None else None,
        'sections':      sections,
    }


# ─── Employee Report (Manager / HR / Super Admin) ─────────────────────────────

def get_employee_report(cycle_id, employee_id, viewer):
    try:
        cycle = ReviewCycle.objects.get(id=cycle_id)
    except ReviewCycle.DoesNotExist:
        raise NotFound('Cycle not found')

    if cycle.state not in ['RESULTS_RELEASED', 'ARCHIVED']:
        raise PermissionDenied('Results are not yet released')

    from apps.users.models import User
    try:
        employee = User.objects.get(id=employee_id)
    except User.DoesNotExist:
        raise NotFound('Employee not found')

    # Manager scope check — can only view direct reports
    if viewer.role == 'MANAGER':
        from apps.users.models import OrgHierarchy
        if not OrgHierarchy.objects.filter(employee=employee, manager=viewer).exists():
            raise PermissionDenied('Employee is not in your team')

    try:
        aggregated = AggregatedResult.objects.get(cycle=cycle, reviewee=employee)
    except AggregatedResult.DoesNotExist:
        aggregated = None

    sections = _get_feedback_sections(cycle, employee, viewer.role, viewer)

    AuditLog.log(actor=viewer, action='VIEW_REPORT', entity_type='report',
                 entity_id=cycle_id, new_value={
                     'viewed_by': viewer.get_full_name(),
                     'viewer_role': viewer.role,
                     'reviewee': employee.get_full_name(),
                     'cycle': cycle.name,
                 })

    return {
        'cycle_id':      str(cycle_id),
        'cycle_name':    cycle.name,
        'reviewee':      {'id': str(employee.id), 'name': employee.get_full_name(), 'email': employee.email},
        'overall_score': float(aggregated.overall_score) if aggregated and aggregated.overall_score is not None else None,
        'self_score':    float(aggregated.self_score)    if aggregated and aggregated.self_score    is not None else None,
        'manager_score': float(aggregated.manager_score) if aggregated and aggregated.manager_score is not None else None,
        'peer_score':    float(aggregated.peer_score)    if aggregated and aggregated.peer_score    is not None else None,
        'sections':      sections,
    }


# ─── Excel Export (Super Admin) ───────────────────────────────────────────────

def export_employee_report_excel(cycle_id, employee_id, actor):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    report = get_employee_report(cycle_id, employee_id, actor)

    from apps.users.models import User
    from apps.review_cycles.models import ReviewCycle
    employee = User.objects.get(id=employee_id)
    cycle    = ReviewCycle.objects.get(id=cycle_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '360 Report'

    header_font  = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    header_fill  = PatternFill('solid', fgColor='1677FF')
    header_font_white = Font(bold=True, color='FFFFFF')

    # ── Title block
    ws.append(['360° Feedback Report'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append(['Cycle',     cycle.name])
    ws.append(['Employee',  employee.get_full_name()])
    ws.append(['Email',     employee.email])
    ws.append(['Job Title', employee.job_title or ''])
    ws.append([])

    # ── Score Summary
    ws.append(['Score Summary'])
    ws[f'A{ws.max_row}'].font = section_font
    ws.append([
        'Overall', report['overall_score'] or '—',
        'Self',    report['self_score']    or '—',
        'Manager', report['manager_score'] or '—',
        'Peer',    report['peer_score']    or '—',
    ])
    ws.append([])

    # ── Detailed Answers header
    header_row = ws.max_row + 1
    ws.append(['Reviewer Type', 'Reviewer', 'Question', 'Rating', 'Text Response'])
    for col in range(1, 6):
        cell = ws.cell(row=header_row, column=col)
        cell.font  = header_font_white
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal='center')

    # ── Rows
    for section in report.get('sections', []):
        if section.get('hidden'):
            reviewer_name = 'Anonymous'
        else:
            identity = section.get('identity') or {}
            reviewer_name = f"{identity.get('first_name', '')} {identity.get('last_name', '')}".strip() or 'Unknown'

        for ans in section.get('answers', []):
            ws.append([
                section['reviewer_type'],
                reviewer_name,
                ans['question_text'],
                ans['rating_value'] if ans['rating_value'] is not None else '',
                ans['text_value'] or '',
            ])

    # ── Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40

    AuditLog.log(actor=actor, action='EXPORT_REPORT', entity_type='report',
                 entity_id=cycle_id, new_value={
                     'exported_by': actor.get_full_name(),
                     'reviewee': employee.get_full_name(),
                     'cycle': cycle.name,
                     'format': 'xlsx',
                 })

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ─── Bulk Excel Export (all employees in a cycle) ─────────────────────────────

def export_all_reports_excel(cycle_id, actor):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    try:
        cycle = ReviewCycle.objects.get(id=cycle_id)
    except ReviewCycle.DoesNotExist:
        raise NotFound('Cycle not found')

    if cycle.state not in ['RESULTS_RELEASED', 'ARCHIVED']:
        raise PermissionDenied('Results are not yet released')

    participants = CycleParticipant.objects.filter(cycle=cycle).select_related('user')

    header_fill       = PatternFill('solid', fgColor='1677FF')
    header_font_white = Font(bold=True, color='FFFFFF', size=11)
    section_font      = Font(bold=True, size=11)
    center            = Alignment(horizontal='center')

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = 'Summary'

    summary_headers = ['Employee', 'Email', 'Department', 'Overall', 'Self', 'Manager', 'Peer']
    ws_sum.append(summary_headers)
    for col, _ in enumerate(summary_headers, 1):
        cell = ws_sum.cell(row=1, column=col)
        cell.font      = header_font_white
        cell.fill      = header_fill
        cell.alignment = center

    for p in participants:
        emp = p.user
        try:
            agg = AggregatedResult.objects.get(cycle=cycle, reviewee=emp)
            overall = round(float(agg.overall_score), 2) if agg.overall_score is not None else '—'
            self_s  = round(float(agg.self_score),    2) if agg.self_score    is not None else '—'
            mgr_s   = round(float(agg.manager_score), 2) if agg.manager_score is not None else '—'
            peer_s  = round(float(agg.peer_score),    2) if agg.peer_score    is not None else '—'
        except AggregatedResult.DoesNotExist:
            overall = self_s = mgr_s = peer_s = '—'

        ws_sum.append([
            emp.get_full_name(),
            emp.email,
            emp.department.name if emp.department else '—',
            overall, self_s, mgr_s, peer_s,
        ])

    ws_sum.column_dimensions['A'].width = 25
    ws_sum.column_dimensions['B'].width = 30
    ws_sum.column_dimensions['C'].width = 20
    for col in ['D', 'E', 'F', 'G']:
        ws_sum.column_dimensions[col].width = 12

    # ── One sheet per employee ────────────────────────────────────────────────
    for p in participants:
        emp = p.user
        sheet_name = emp.get_full_name()[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)

        # Title block
        ws.append(['360° Feedback Report'])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append(['Cycle',     cycle.name])
        ws.append(['Employee',  emp.get_full_name()])
        ws.append(['Email',     emp.email])
        ws.append(['Job Title', emp.job_title or '—'])
        ws.append([])

        # Scores
        ws.append(['Score Summary'])
        ws.cell(row=ws.max_row, column=1).font = section_font
        try:
            agg = AggregatedResult.objects.get(cycle=cycle, reviewee=emp)
            ws.append([
                'Overall', round(float(agg.overall_score), 2) if agg.overall_score is not None else '—',
                'Self',    round(float(agg.self_score),    2) if agg.self_score    is not None else '—',
                'Manager', round(float(agg.manager_score), 2) if agg.manager_score is not None else '—',
                'Peer',    round(float(agg.peer_score),    2) if agg.peer_score    is not None else '—',
            ])
        except AggregatedResult.DoesNotExist:
            ws.append(['No aggregated scores available'])
        ws.append([])

        # Detailed answers header
        header_row = ws.max_row + 1
        ws.append(['Reviewer Type', 'Reviewer', 'Question', 'Rating', 'Text Response'])
        for col in range(1, 6):
            cell = ws.cell(row=header_row, column=col)
            cell.font      = header_font_white
            cell.fill      = header_fill
            cell.alignment = center

        # Answers (HR always sees identity)
        sections = _get_feedback_sections(cycle, emp, actor.role, actor)
        for section in sections:
            if section.get('hidden'):
                reviewer_name = 'Anonymous'
            else:
                identity = section.get('identity') or {}
                reviewer_name = f"{identity.get('first_name', '')} {identity.get('last_name', '')}".strip() or 'Unknown'

            for ans in section.get('answers', []):
                ws.append([
                    section['reviewer_type'],
                    reviewer_name,
                    ans['question_text'],
                    ans['rating_value'] if ans['rating_value'] is not None else '',
                    ans['text_value'] or '',
                ])

        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 40

    AuditLog.log(actor=actor, action='EXPORT_REPORT', entity_type='report',
                 entity_id=cycle_id, new_value={
                     'exported_by': actor.get_full_name(),
                     'cycle': cycle.name,
                     'format': 'xlsx_bulk',
                     'participant_count': participants.count(),
                 })

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
